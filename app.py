import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from datetime import datetime
import requests
import io

WSD_URL  = "https://docs.google.com/spreadsheets/d/1zfRWlDKS--rTqD4qcK2VJFgaQE2iQ8vu/export?format=xlsx"
IMHS_URL = "https://docs.google.com/spreadsheets/d/1T3nl9w5DpT3yOVVQOhzZpamA0qVGQA_n/export?format=xlsx"
ZCHS_URL = "https://docs.google.com/spreadsheets/d/1sqyx9DUO3Dhbe51hPP5YRTNhCutY1b6_/export?format=xlsx"

def fetch_wb(url):
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

st.set_page_config(
    page_title="Worldsprings Pricing Dashboard",
    page_icon="♨️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── password gate ─────────────────────────────────────────────────────────────
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.markdown("""
    <div style='max-width:360px;margin:120px auto 0;text-align:center'>
        <div style='font-size:40px;margin-bottom:12px'>♨️</div>
        <div style='font-size:24px;font-weight:700;color:#F8FAFC;margin-bottom:4px'>Worldsprings</div>
        <div style='font-size:13px;color:#64748B;margin-bottom:32px'>Pricing Dashboard</div>
    </div>
    """, unsafe_allow_html=True)
    col_l, col_m, col_r = st.columns([1, 1.2, 1])
    with col_m:
        pwd = st.text_input("Password", type="password", placeholder="Enter password")
        if st.button("Enter", use_container_width=True):
            if pwd == st.secrets["password"]:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password")
    st.stop()

# ── color palette ─────────────────────────────────────────────────────────────
C = {
    "WSD":  {"p": "#38BDF8", "light": "#BAE6FD", "dark": "#0369A1", "bg": "#051525", "glow": "#38BDF820"},
    "IMHS": {"p": "#4ADE80", "light": "#BBF7D0", "dark": "#15803D", "bg": "#051505", "glow": "#4ADE8020"},
    "ZCHS": {"p": "#F87171", "light": "#FECACA", "dark": "#B91C1C", "bg": "#1a0505", "glow": "#F8717120"},
    "base": "#0D1117",
}

DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# ── base CSS ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

* { font-family: 'Inter', sans-serif !important; }

/* App background */
.stApp { background-color: #0D1117; }
[data-testid="stAppViewContainer"] { background-color: #0D1117; }

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #090D13 0%, #0D1117 100%);
    border-right: 1px solid #1E2D3D;
    min-width: 168px !important;
    max-width: 168px !important;
    width: 168px !important;
}
[data-testid="stSidebar"] * { color: #94A3B8 !important; }
[data-testid="stSidebar"] .stRadio label { font-size: 13px !important; }

/* Radio buttons in sidebar */
[data-testid="stSidebar"] [data-testid="stWidgetLabel"] { display: none; }

/* Tabs */
button[data-baseweb="tab"] {
    background: transparent !important;
    color: #94A3B8 !important;
    border-bottom: 2px solid transparent !important;
    font-weight: 500 !important;
    font-size: 14px !important;
}
button[data-baseweb="tab"]:hover {
    color: #F8FAFC !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color: #FFFFFF !important;
    font-weight: 600 !important;
    border-bottom: 2px solid currentColor !important;
}

/* DataFrames */
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }

/* Scrollbar */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0D1117; }
::-webkit-scrollbar-thumb { background: #1E2D3D; border-radius: 3px; }
</style>
""", unsafe_allow_html=True)


# ── helpers ───────────────────────────────────────────────────────────────────
def inject_page_bg(loc: str):
    bg = C[loc]["bg"]
    glow = C[loc]["glow"]
    st.markdown(f"""
    <style>
    .stApp, [data-testid="stAppViewContainer"] {{
        background: radial-gradient(ellipse at top left, {glow} 0%, {bg} 40%, #0D1117 100%) !important;
    }}
    </style>
    """, unsafe_allow_html=True)


def banner(title: str, subtitle: str, loc: str, emoji: str):
    p = C[loc]["p"]
    light = C[loc]["light"]
    dark = C[loc]["dark"]
    st.markdown(f"""
    <div style='
        background: linear-gradient(135deg, {dark}55 0%, {p}22 100%);
        border: 1px solid {p}44;
        border-radius: 16px;
        padding: 28px 32px;
        margin-bottom: 28px;
        position: relative;
        overflow: hidden;
    '>
        <div style='font-size:36px;margin-bottom:8px'>{emoji}</div>
        <div style='font-size:28px;font-weight:700;color:{light};letter-spacing:-0.5px'>{title}</div>
        <div style='font-size:14px;color:{p};margin-top:6px;font-weight:500'>{subtitle}</div>
        <div style='
            position:absolute;right:-20px;top:-20px;
            width:120px;height:120px;
            background:{p}15;
            border-radius:50%;
        '></div>
    </div>
    """, unsafe_allow_html=True)


def section_card(content_html: str, loc: str):
    p = C[loc]["p"]
    st.markdown(f"""
    <div style='
        background: linear-gradient(135deg, #ffffff08, #ffffff04);
        border: 1px solid {p}25;
        border-radius: 12px;
        padding: 20px 22px;
        margin-bottom: 16px;
    '>{content_html}</div>
    """, unsafe_allow_html=True)


def price_table(data: dict, title: str, loc: str, subtitle: str = "", compact: bool = False):
    p = C[loc]["p"]
    light = C[loc]["light"]
    dark = C[loc]["dark"]
    hp = "5px 7px" if compact else "8px 16px"
    cp = "6px 7px" if compact else "9px 16px"
    hfs = "11px" if compact else "12px"
    vfs = "12px" if compact else "15px"
    lfs = "11px" if compact else "13px"

    header_cols = "".join(
        f"<th style='padding:{hp};background:{dark}55;color:{light};font-weight:600;font-size:{hfs};letter-spacing:0.3px;text-align:center'>{h}</th>"
        for h in next(iter(data.values())).keys()
    )

    rows_html = ""
    for i, (product, prices) in enumerate(data.items()):
        bg = "#ffffff06" if i % 2 == 0 else "transparent"
        cols = ""
        for v in prices.values():
            if isinstance(v, (int, float)):
                display = int(v) if isinstance(v, float) and v == int(v) else v
                cols += f"<td style='padding:{cp};text-align:center;font-weight:600;color:#F8FAFC;font-size:{vfs}'>${display}</td>"
            else:
                cols += f"<td style='padding:{cp};text-align:center;color:#64748B;font-size:{lfs}'>{v if v else '—'}</td>"
        rows_html += f"<tr style='background:{bg}'><td style='padding:{cp};font-weight:500;color:#CBD5E1;font-size:{lfs}'>{product}</td>{cols}</tr>"

    sub = f"<div style='font-size:12px;color:#64748B;margin-top:2px'>{subtitle}</div>" if subtitle else ""
    table = f"""
    <div style='margin-bottom:20px'>
      <div style='font-size:13px;font-weight:700;color:{p};text-transform:uppercase;letter-spacing:1px;margin-bottom:4px'>{title}</div>
      {sub}
      <table style='width:100%;border-collapse:collapse;margin-top:10px;font-size:13px;'>
        <thead>
          <tr>
            <th style='padding:8px 16px;background:{dark}55;color:{light};font-weight:600;font-size:12px;letter-spacing:0.5px;text-align:left'>Product</th>
            {header_cols}
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>"""
    st.markdown(table, unsafe_allow_html=True)


def simple_table(rows: list, title: str, loc: str):
    """rows = list of (label, value) tuples"""
    p = C[loc]["p"]
    rows_html = "".join(
        f"<tr style='background:{'#ffffff06' if i%2==0 else 'transparent'}'>"
        f"<td style='padding:8px 14px;color:#CBD5E1;font-size:13px'>{r[0]}</td>"
        f"<td style='padding:8px 14px;text-align:right;font-weight:600;color:#F8FAFC;font-size:14px'>{r[1]}</td></tr>"
        for i, r in enumerate(rows)
    )
    st.markdown(f"""
    <div style='margin-bottom:20px'>
      <div style='font-size:13px;font-weight:700;color:{p};text-transform:uppercase;letter-spacing:1px;margin-bottom:8px'>{title}</div>
      <table style='width:100%;border-collapse:collapse;'>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    """, unsafe_allow_html=True)


def chart_layout(fig, _loc: str, height: int = 420):
    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#FFFFFF", family="Inter"),
        height=height,
        legend=dict(
            orientation="h", y=-0.22,
            bgcolor="rgba(0,0,0,0)",
            font=dict(color="#FFFFFF", size=12),
        ),
        yaxis=dict(
            title=dict(text="Price (USD)", font=dict(color="#FFFFFF")),
            gridcolor="#1E2D3D",
            tickfont=dict(color="#FFFFFF"),
            zeroline=False,
        ),
        xaxis=dict(
            tickfont=dict(color="#FFFFFF"),
        ),
        margin=dict(t=20, b=60, l=10, r=10),
    )


# ── data loaders ──────────────────────────────────────────────────────────────
@st.cache_data
def load_wsd():
    wb = fetch_wb(WSD_URL)
    ws = wb["2026 rates "]
    rows = list(ws.iter_rows(values_only=True))

    current = {
        "Day Pass":     {"Mon-Thu": 85, "Fri & Sun": 95, "Saturday": 105},
        "Evening Pass": {"Mon-Thu": 55, "Fri & Sun": 65, "Saturday": 85},
    }

    spa_current = []
    for row in rows[1:]:
        if row[0] == "Spa " and row[1] and row[2] is not None:
            try:
                spa_current.append({
                    "Service": row[1],
                    "Mon-Thu": row[2] if isinstance(row[2], (int, float)) else None,
                    "Fri & Sun": row[3] if isinstance(row[3], (int, float)) else None,
                    "Saturday": row[4] if isinstance(row[4], (int, float)) else None,
                })
            except Exception:
                pass

    history = []
    for label, prod, mth, frs in [
        ("Jul-Sep 2024", "3 Hour Soak", 55, 55),
        ("Oct-Jan 2025", "3 Hour Soak", 59, 85),
        ("Feb-May 2025", "3 Hour Soak", 49, 69),
        ("Jun-Dec 2025", "3 Hour Soak", 69, 79),
        ("Jan-Feb 2026", "Day Pass",    85, 95),
    ]:
        history.append({"Period": label, "Product": prod, "Mon-Thu": mth, "Fri-Sun": frs})
    for label, mth, frs in [
        ("Jul-Sep 2024", 35, 35), ("Oct-Jan 2025", 39, 39),
        ("Feb-May 2025", 39, 39), ("Jun-Dec 2025", 49, 59),
    ]:
        history.append({"Period": label, "Product": "Quick Dip", "Mon-Thu": mth, "Fri-Sun": frs})
    history.append({"Period": "Jan-Feb 2026", "Product": "Evening Pass", "Mon-Thu": 55, "Fri-Sun": 65})

    return current, pd.DataFrame(history), pd.DataFrame(spa_current)


@st.cache_data
def load_imhs():
    wb = fetch_wb(IMHS_URL)
    ws = wb["Pricing 3.30.26-5.31.26"]
    rows = list(ws.iter_rows(values_only=True))
    non_peak, peak, section = {}, {}, None
    for row in rows:
        label = row[2] if len(row) > 2 else None
        if label == "Non-Peak Pricing":
            section = "non_peak"
        elif label == "Holiday/Peak Pricing":
            section = "peak"
        elif label in DAY_ORDER and section:
            d = {"Select 3hr": row[3], "Select All-Day": row[4],
                 "Premier 3hr": row[5], "Premier All-Day": row[6]}
            (non_peak if section == "non_peak" else peak)[label] = d

    # ── Hardcoded price history (1/23/23 – 5/31/26) ───────────────────────────
    # Each entry: (period_label, date_str, non_peak_map, peak_map)
    # Day map values: (Select 3hr, Select All-Day, Premier 3hr, Premier All-Day)
    # None = product didn't exist in that period
    _NP = "Non-Peak"
    _PK = "Peak"

    HIST_PERIODS = [
        # 1/23/23–4/30/23: single unified ticket (≈ Premier access), no Select/Peak split
        ("1.23.23", "1.23.2023",
         {"Mon":(None,None,40,100),"Tue":(None,None,40,100),"Wed":(None,None,40,100),
          "Thu":(None,None,40,100),"Fri":(None,None,44,None),
          "Sat":(None,None,48,125),"Sun":(None,None,48,125)},
         None),

        # 5/1/23–9/4/23: GA (Select) + WS (Premier) introduced, no Peak
        ("5.1.23", "5.1.2023",
         {"Mon":(39,None,49,100),"Tue":(39,None,49,100),"Wed":(39,None,49,100),
          "Thu":(39,None,49,100),"Fri":(44,None,54,150),
          "Sat":(48,None,58,150),"Sun":(48,None,58,150)},
         None),

        # 9/5/23–10/31/23
        ("9.5.23", "9.5.2023",
         {"Mon":(40,None,50,100),"Tue":(40,None,50,100),"Wed":(40,None,50,100),
          "Thu":(40,None,50,100),"Fri":(48,None,58,150),
          "Sat":(48,None,58,150),"Sun":(48,None,58,150)},
         None),

        # 11/1/23–12/21/23
        ("11.1.23", "11.1.2023",
         {"Mon":(40,None,50,100),"Tue":(40,None,50,100),"Wed":(40,None,50,100),
          "Thu":(40,None,50,100),"Fri":(48,None,58,150),
          "Sat":(48,None,58,150),"Sun":(48,None,58,150)},
         None),

        # 12/22/23–5/23/24: +$4 GA / +$9 WS; Peak pricing introduced
        ("12.22.23", "12.22.2023",
         {"Mon":(40,None,59,100),"Tue":(40,None,59,100),"Wed":(40,None,59,100),
          "Thu":(40,None,59,100),"Fri":(52,None,67,150),
          "Sat":(52,None,67,150),"Sun":(52,None,67,150)},
         {"Mon":(52,None,67,150),"Tue":(52,None,67,150),"Wed":(52,None,67,150),
          "Thu":(52,None,67,150),"Fri":(52,None,67,150),
          "Sat":(52,None,67,150),"Sun":(52,None,67,150)}),

        # 5/24/24–11/3/24
        ("5.24.24", "5.24.2024",
         {"Mon":(40,None,59,100),"Tue":(40,None,59,100),"Wed":(40,None,59,100),
          "Thu":(40,None,59,100),"Fri":(52,None,67,150),
          "Sat":(52,None,67,150),"Sun":(52,None,67,150)},
         {"Mon":(52,None,67,150),"Tue":(52,None,67,150),"Wed":(52,None,67,150),
          "Thu":(52,None,67,150),"Fri":(52,None,67,150),
          "Sat":(52,None,67,150),"Sun":(52,None,67,150)}),

        # 11/4/24–2/28/25
        ("11.4.24", "11.4.2024",
         {"Mon":(44,None,59,100),"Tue":(44,None,59,100),"Wed":(44,None,59,100),
          "Thu":(44,None,59,100),"Fri":(52,None,67,150),
          "Sat":(52,None,67,150),"Sun":(52,None,67,150)},
         {"Mon":(52,None,67,150),"Tue":(52,None,67,150),"Wed":(52,None,67,150),
          "Thu":(52,None,67,150),"Fri":(52,None,67,150),
          "Sat":(52,None,67,150),"Sun":(52,None,67,150)}),

        # 3/1/25–6/30/25: timeslot-era pricing (GA/WS), no Peak
        ("3.1.25", "3.1.2025",
         {"Mon":(46,None,63,100),"Tue":(46,None,63,100),"Wed":(46,None,63,100),
          "Thu":(46,None,63,100),"Fri":(48,None,65,150),
          "Sat":(48,None,65,150),"Sun":(52,None,69,150)},
         None),

        # 7/1/25–8/30/25: same as above
        ("7.1.25", "7.1.2025",
         {"Mon":(46,None,63,100),"Tue":(46,None,63,100),"Wed":(46,None,63,100),
          "Thu":(46,None,63,100),"Fri":(48,None,65,150),
          "Sat":(48,None,65,150),"Sun":(52,None,69,150)},
         None),

        # 9/1/25–9/30/25: new structured format — Select All-Day introduced
        ("9.1.25", "9.1.2025",
         {"Mon":(48,53,68,78),"Tue":(48,53,68,78),"Wed":(48,53,68,78),
          "Thu":(48,53,68,78),"Fri":(48,58,68,83),
          "Sat":(56,76,76,116),"Sun":(56,66,76,91)},
         {"Mon":(56,76,76,116)}),  # Labor Day only

        # 10/1/25–2/1/26
        ("10.1.25", "10.1.2025",
         {"Mon":(48,53,68,78),"Tue":(48,53,68,78),"Wed":(48,53,68,78),
          "Thu":(48,53,68,78),"Fri":(48,58,68,83),
          "Sat":(56,76,76,116),"Sun":(56,66,76,91)},
         {"Mon":(56,96,76,136),"Tue":(56,96,76,136),"Wed":(56,96,76,136),
          "Thu":(56,96,76,136),"Fri":(56,96,76,136),
          "Sat":(56,96,76,136),"Sun":(56,96,76,136)}),

        # 2/2/26–3/29/26 (standard pricing)
        ("2.2.26", "2.2.2026",
         {"Mon":(48,53,68,78),"Tue":(48,53,68,78),"Wed":(48,53,68,78),
          "Thu":(48,53,68,78),"Fri":(48,58,68,83),
          "Sat":(56,76,76,150),"Sun":(56,66,76,130)},
         {"Mon":(60,75,95,150),"Tue":(60,75,95,150),"Wed":(60,75,95,150),
          "Thu":(60,75,95,150),"Fri":(60,75,95,150),
          "Sat":(60,75,95,150),"Sun":(60,75,95,150)}),

        # 3/2/26: Summit week (special pricing within 2/2/26–3/29/26)
        ("3.2.26", "3.2.2026",
         {"Mon":(50,55,75,80),"Tue":(50,55,75,80),"Wed":(50,55,75,80),
          "Thu":(50,55,75,80),"Fri":(50,60,80,90),
          "Sat":(60,75,95,150),"Sun":(55,65,85,130)},
         {"Mon":(60,75,95,150),"Tue":(60,75,95,150),"Wed":(60,75,95,150),
          "Thu":(60,75,95,150),"Fri":(60,75,95,150),
          "Sat":(60,75,95,150),"Sun":(60,75,95,150)}),

        # 3/30/26–5/31/26
        ("3.30.26", "3.30.2026",
         {"Mon":(50,55,75,80),"Tue":(50,55,75,80),"Wed":(50,55,75,80),
          "Thu":(50,55,75,80),"Fri":(50,60,80,90),
          "Sat":(60,75,95,150),"Sun":(55,65,85,130)},
         {"Mon":(60,75,95,150),"Tue":(60,75,95,150),"Wed":(60,75,95,150),
          "Thu":(60,75,95,150),"Fri":(60,75,95,150),
          "Sat":(60,75,95,150),"Sun":(60,75,95,150)}),
    ]

    hist_records = []
    for label, date_str, np_map, pk_map in HIST_PERIODS:
        for type_, dmap in [(_NP, np_map), (_PK, pk_map)]:
            if not dmap:
                continue
            for day in DAY_ORDER:
                if day not in dmap:
                    continue
                s3, sad, p3, pad = dmap[day]
                hist_records.append({
                    "Period": label, "Date": date_str, "Type": type_, "Day": day,
                    "Select 3hr": s3, "Select All-Day": sad,
                    "Premier 3hr": p3, "Premier All-Day": pad,
                })

    return non_peak, peak, pd.DataFrame(hist_records)


@st.cache_data
def load_zchs():
    wb = fetch_wb(ZCHS_URL)
    ws = wb["Pricing 11_2025"]
    rows = list(ws.iter_rows(values_only=True))
    current = {}
    for row in rows:
        label = str(row[0]).strip() if row[0] else ""
        if "Quick Dip Adult / Teen" in label and "Select" in label:
            if "Quick Dip Select (13+)" not in current:
                current["Quick Dip Select (13+)"] = {"Mon-Thu": row[1], "Fri/Sat/Sun": row[2]}
        elif "Quick Dip Adult: Premier" in label:
            if "Quick Dip Premier (21+)" not in current:
                current["Quick Dip Premier (21+)"] = {"Mon-Thu": row[1], "Fri/Sat/Sun": row[2]}
        elif "3-hour Adult / Teen" in label and "Select" in label and "Eat" not in label:
            if "3hr Select (13+)" not in current:
                current["3hr Select (13+)"] = {"Mon-Thu": row[1], "Fri/Sat/Sun": row[2]}
        elif "3- Hour Soak Adult: Premier" in label:
            if "3hr Premier (21+)" not in current:
                current["3hr Premier (21+)"] = {"Mon-Thu": row[1], "Fri/Sat/Sun": row[2]}
        elif "Full Day Soak Ages 13+" in label:
            if "Full Day Select (13+)" not in current:
                current["Full Day Select (13+)"] = {"Mon-Thu": row[1], "Fri/Sat/Sun": row[2]}
        elif "Full Day Adult: Premier" in label:
            if "Full Day Premier (21+)" not in current:
                current["Full Day Premier (21+)"] = {"Mon-Thu": row[1], "Fri/Sat/Sun": row[2]}

    pricing_sheets = {
        "Feb 2025": ("Pricing 2_3_2025", "2.3.2025"),
        "Jul 2025": ("Pricing 07_2025",  "7.1.2025"),
        "Nov 2025": ("Pricing 11_2025",  "11.1.2025"),
    }
    hist_records = []
    seen = {}
    for period, (sheet_name, date_s) in pricing_sheets.items():
        if sheet_name not in wb.sheetnames: continue
        seen[period] = set()
        in_section = None  # tracks current section for Jan 2025 format

        for row in wb[sheet_name].iter_rows(values_only=True):
            lbl = str(row[0]).strip() if row[0] else ""
            v1 = row[1] if isinstance(row[1], (int, float)) else None
            v2 = row[2] if isinstance(row[2], (int, float)) else None

            # Detect section headers (non-numeric col B = header row)
            if v1 is None:
                if "Quick Dip" in lbl:
                    in_section = "qd"
                elif "Day Pass" in lbl or ("Full Day" in lbl and "Soak" in lbl):
                    in_section = "fd"
                continue

            def _add(prod):
                key = (prod, period)
                if key not in seen[period]:
                    seen[period].add(key)
                    hist_records.append({"Period": period, "Date": date_s,
                                         "Product": prod, "Mon-Thu": v1, "Fri/Sat/Sun": v2})

            # Jan 2025 format: bare "Select Access" row identified by section context
            # "Quick Dip (3 hours)" in Jan 2025 = same 3-hour product as "3hr Select" later
            # The short Quick Dip (~1.5hr) is a new product added from Jul 2025 onward
            if lbl == "Select Access":
                if in_section == "qd":
                    _add("3hr Select (13+)")
                elif in_section == "fd":
                    _add("Full Day Select (13+)")

            # Jul/Nov 2025 format: full descriptive labels
            if "Quick Dip Adult / Teen" in lbl and "Select" in lbl:
                _add("Quick Dip Select (13+)")
            elif "3-hour Adult / Teen" in lbl and "Select" in lbl and "Eat" not in lbl:
                _add("3hr Select (13+)")
            elif "Full Day Soak Ages 13+" in lbl:
                _add("Full Day Select (13+)")

    return current, pd.DataFrame(hist_records)


# ── sidebar nav ───────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style='padding: 20px 8px 8px 8px;'>
        <div style='font-size:17px;font-weight:700;color:#F8FAFC;letter-spacing:-0.5px'>2026 Price</div>
        <div style='font-size:17px;font-weight:700;color:#F8FAFC;letter-spacing:-0.5px'>Dashboard</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    page = st.radio(
        "nav",
        ["🌐  Overview", "🔵  WSD – Dallas", "🟢  IMHS – Colorado", "🔴  ZCHS – Utah"],
        label_visibility="collapsed",
    )

    st.markdown(f"<div style='margin-top:24px;font-size:11px;color:#334155;border-top:1px solid #1E2D3D;padding-top:10px'>Updated {datetime.today().strftime('%b %d, %Y')}</div>", unsafe_allow_html=True)


def base_chart():
    return dict(
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#FFFFFF", family="Inter"),
        legend=dict(orientation="h", y=-0.22, bgcolor="rgba(0,0,0,0)", font=dict(color="#FFFFFF", size=12)),
        yaxis=dict(gridcolor="#1E2D3D", tickfont=dict(color="#FFFFFF"), zeroline=False,
                   title=dict(text="Price (USD)", font=dict(color="#FFFFFF"))),
        xaxis=dict(tickfont=dict(color="#FFFFFF")),
        margin=dict(t=20, b=60, l=10, r=10),
    )


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
if "Overview" in page:
    wsd_current, _, _ = load_wsd()
    imhs_non_peak, imhs_peak, _ = load_imhs()
    zchs_current, _ = load_zchs()

    st.markdown("""
    <div style='padding:32px 0 16px 0'>
        <div style='font-size:32px;font-weight:700;color:#F8FAFC;letter-spacing:-1px'>Pricing Overview</div>
        <div style='font-size:14px;color:#64748B;margin-top:6px'>All three locations — current rates at a glance</div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.7, 1.3])

    with col1:
        wsd_p = C["WSD"]["p"]; wsd_d = C["WSD"]["dark"]
        st.markdown(f"""<div style='background:linear-gradient(135deg,{wsd_d}55,{wsd_p}20);
            border:1px solid {wsd_p}40;border-radius:12px;padding:14px 18px;margin-bottom:12px'>
            <span style='font-size:11px;font-weight:700;color:{wsd_p};text-transform:uppercase;letter-spacing:1.5px'>WSD</span>
            <span style='font-size:15px;font-weight:700;color:#F8FAFC;margin-left:10px'>Dallas, TX</span>
        </div>""", unsafe_allow_html=True)
        price_table(wsd_current, "Soak Admission", "WSD", compact=True)
        simple_table([("1-Month Trial","$125"),("Weekday Annual","$948/yr"),("Anytime Annual","$1,788/yr")], "Memberships", "WSD")
        simple_table([("Robe Rental","$10"),("Cabana Mon-Thu","$199"),("Cabana Sat","$299")], "Add-ons", "WSD")

    with col2:
        imhs_p = C["IMHS"]["p"]; imhs_d = C["IMHS"]["dark"]
        st.markdown(f"""<div style='background:linear-gradient(135deg,{imhs_d}55,{imhs_p}20);
            border:1px solid {imhs_p}40;border-radius:12px;padding:14px 18px;margin-bottom:12px'>
            <span style='font-size:11px;font-weight:700;color:{imhs_p};text-transform:uppercase;letter-spacing:1.5px'>IMHS</span>
            <span style='font-size:15px;font-weight:700;color:#F8FAFC;margin-left:10px'>Glenwood Springs, CO</span>
        </div>""", unsafe_allow_html=True)
        np_t = {d: {"Sel 3hr": imhs_non_peak.get(d,{}).get("Select 3hr"),
                    "Sel All-Day": imhs_non_peak.get(d,{}).get("Select All-Day"),
                    "Pre 3hr": imhs_non_peak.get(d,{}).get("Premier 3hr"),
                    "Pre All-Day": imhs_non_peak.get(d,{}).get("Premier All-Day")}
                for d in DAY_ORDER if d in imhs_non_peak}
        price_table(np_t, "Non-Peak Pricing", "IMHS", compact=True)
        pk_t = {d: {"Sel 3hr": imhs_peak.get(d,{}).get("Select 3hr"),
                    "Sel All-Day": imhs_peak.get(d,{}).get("Select All-Day"),
                    "Pre 3hr": imhs_peak.get(d,{}).get("Premier 3hr"),
                    "Pre All-Day": imhs_peak.get(d,{}).get("Premier All-Day")}
                for d in DAY_ORDER if d in imhs_peak}
        price_table(pk_t, "Peak / Holiday Pricing", "IMHS", compact=True)

    with col3:
        zchs_p = C["ZCHS"]["p"]; zchs_d = C["ZCHS"]["dark"]
        st.markdown(f"""<div style='background:linear-gradient(135deg,{zchs_d}55,{zchs_p}20);
            border:1px solid {zchs_p}40;border-radius:12px;padding:14px 18px;margin-bottom:12px'>
            <span style='font-size:11px;font-weight:700;color:{zchs_p};text-transform:uppercase;letter-spacing:1.5px'>ZCHS</span>
            <span style='font-size:15px;font-weight:700;color:#F8FAFC;margin-left:10px'>Zion Canyon, UT</span>
        </div>""", unsafe_allow_html=True)
        if zchs_current:
            disp = {pr: {k: v if isinstance(v,(int,float)) else "—" for k,v in prc.items()} for pr,prc in zchs_current.items()}
            price_table(disp, "Soak Pricing (May 2026)", "ZCHS", compact=True)
        simple_table([("Anytime Annual","$1,499"),("Weekday Annual","$920"),("Snowbird (3 mo)","$499")], "Memberships", "ZCHS")
        simple_table([("Single Cabana Mon-Thu","$149"),("Single Cabana Fri-Sun","$249"),
                      ("Double Cabana Mon-Thu","$199"),("Double Cabana Fri-Sun","$299")], "Add-ons", "ZCHS")

    st.markdown("<div style='margin-top:32px'></div>", unsafe_allow_html=True)
    st.markdown("""<div style='font-size:18px;font-weight:700;color:#F8FAFC;margin-bottom:4px'>Comparable Entry Price</div>
    <div style='font-size:13px;color:#64748B;margin-bottom:16px'>Standard soak — weekday vs weekend</div>""", unsafe_allow_html=True)

    wsd_p=C["WSD"]["p"]; imhs_p=C["IMHS"]["p"]; zchs_p=C["ZCHS"]["p"]
    wsd_d=C["WSD"]["dark"]; imhs_d=C["IMHS"]["dark"]; zchs_d=C["ZCHS"]["dark"]

    fig = go.Figure()
    fig.add_trace(go.Bar(name="WSD Weekday", x=["WSD Dallas"], y=[85], marker_color=wsd_p,
        text=["$85"], textposition="outside", textfont=dict(color=wsd_p)))
    fig.add_trace(go.Bar(name="WSD Weekend", x=["WSD Dallas"], y=[105], marker_color=wsd_d,
        text=["$105"], textposition="outside", textfont=dict(color=wsd_p)))
    fig.add_trace(go.Bar(name="IMHS Weekday", x=["IMHS Colorado"],
        y=[imhs_non_peak.get("Mon",{}).get("Select All-Day",0)], marker_color=imhs_p,
        text=[f"${imhs_non_peak.get('Mon',{}).get('Select All-Day',0)}"],
        textposition="outside", textfont=dict(color=imhs_p)))
    fig.add_trace(go.Bar(name="IMHS Weekend Peak", x=["IMHS Colorado"],
        y=[imhs_peak.get("Sat",{}).get("Premier All-Day",0)], marker_color=imhs_d,
        text=[f"${imhs_peak.get('Sat',{}).get('Premier All-Day',0)}"],
        textposition="outside", textfont=dict(color=imhs_p)))
    fig.add_trace(go.Bar(name="ZCHS Weekday", x=["ZCHS Utah"],
        y=[zchs_current.get("3hr Select (13+)",{}).get("Mon-Thu",0)], marker_color=zchs_p,
        text=[f"${zchs_current.get('3hr Select (13+)',{}).get('Mon-Thu',0)}"],
        textposition="outside", textfont=dict(color=zchs_p)))
    fig.add_trace(go.Bar(name="ZCHS Weekend", x=["ZCHS Utah"],
        y=[zchs_current.get("3hr Select (13+)",{}).get("Fri/Sat/Sun",0)], marker_color=zchs_d,
        text=[f"${zchs_current.get('3hr Select (13+)',{}).get('Fri/Sat/Sun',0)}"],
        textposition="outside", textfont=dict(color=zchs_p)))

    layout = base_chart()
    layout.update(barmode="group", height=360,
        xaxis=dict(tickfont=dict(color="#FFFFFF", size=13)),
        bargap=0.3, bargroupgap=0.08)
    fig.update_layout(**layout)
    st.plotly_chart(fig, use_container_width=True)

    # ── price by day filter ───────────────────────────────────────────────────
    st.markdown("""<div style='font-size:18px;font-weight:700;color:#F8FAFC;margin-bottom:4px'>Price by Day of Week</div>
    <div style='font-size:13px;color:#64748B;margin-bottom:16px'>Select a day to compare admission prices across the three locations</div>""",
    unsafe_allow_html=True)

    selected_day = st.radio("Day", DAY_ORDER, horizontal=True, label_visibility="collapsed")

    # WSD: map day → tier
    wsd_tier_map = {"Mon":"Mon-Thu","Tue":"Mon-Thu","Wed":"Mon-Thu","Thu":"Mon-Thu",
                    "Fri":"Fri & Sun","Sat":"Saturday","Sun":"Fri & Sun"}
    wsd_day_price  = wsd_current["Day Pass"][wsd_tier_map[selected_day]]
    wsd_eve_price  = wsd_current["Evening Pass"][wsd_tier_map[selected_day]]

    # IMHS: non_peak prices for that day
    imhs_sel3  = imhs_non_peak.get(selected_day, {}).get("Select 3hr", 0)
    imhs_selAD = imhs_non_peak.get(selected_day, {}).get("Select All-Day", 0)
    imhs_pre3  = imhs_non_peak.get(selected_day, {}).get("Premier 3hr", 0)
    imhs_preAD = imhs_non_peak.get(selected_day, {}).get("Premier All-Day", 0)

    # ZCHS: Mon-Thu vs Fri/Sat/Sun
    zchs_key = "Mon-Thu" if selected_day in ("Mon","Tue","Wed","Thu") else "Fri/Sat/Sun"
    zchs_qd  = zchs_current.get("Quick Dip Select (13+)", {}).get(zchs_key, 0)
    zchs_3hr = zchs_current.get("3hr Select (13+)", {}).get(zchs_key, 0)
    zchs_fd  = zchs_current.get("Full Day Select (13+)", {}).get(zchs_key, 0)

    wsd_p=C["WSD"]["p"]; imhs_p=C["IMHS"]["p"]; zchs_p=C["ZCHS"]["p"]
    wsd_d=C["WSD"]["dark"]; imhs_d=C["IMHS"]["dark"]; zchs_d=C["ZCHS"]["dark"]

    # subplots: one per location so bars are always centered
    fig2 = make_subplots(rows=1, cols=3,
        subplot_titles=["WSD — Dallas", "IMHS — Colorado", "ZCHS — Utah"],
        shared_yaxes=True)

    # WSD — Day Pass + Evening Pass
    for label, val, color in [
        ("Day Pass",    wsd_day_price, wsd_p),
        ("Evening Pass", wsd_eve_price, wsd_d),
    ]:
        fig2.add_trace(go.Bar(name=label, x=[label], y=[val], marker_color=color,
            text=[f"${val}"], textposition="outside", textfont=dict(color="#FFFFFF"),
            showlegend=False), row=1, col=1)

    # IMHS — 4 tiers
    for label, val, color in [
        ("Sel 3hr",  imhs_sel3,  imhs_p),
        ("Sel All-Day", imhs_selAD, "#22c55e"),
        ("Pre 3hr",  imhs_pre3,  "#86efac"),
        ("Pre All-Day", imhs_preAD, imhs_d),
    ]:
        fig2.add_trace(go.Bar(name=label, x=[label], y=[val], marker_color=color,
            text=[f"${val}"], textposition="outside", textfont=dict(color="#FFFFFF"),
            showlegend=False), row=1, col=2)

    # ZCHS — 3 products
    for label, val, color in [
        ("Quick Dip", zchs_qd,  zchs_p),
        ("3hr Select", zchs_3hr, "#ef4444"),
        ("Full Day",  zchs_fd,  zchs_d),
    ]:
        fig2.add_trace(go.Bar(name=label, x=[label], y=[val], marker_color=color,
            text=[f"${val}"], textposition="outside", textfont=dict(color="#FFFFFF"),
            showlegend=False), row=1, col=3)

    fig2.update_layout(
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#FFFFFF", family="Inter"),
        height=400, showlegend=False,
        margin=dict(t=40, b=20, l=10, r=10),
    )
    fig2.update_annotations(font=dict(color="#FFFFFF", size=13))
    for i in range(1, 4):
        fig2.update_xaxes(tickfont=dict(color="#FFFFFF", size=11), row=1, col=i)
        fig2.update_yaxes(gridcolor="#1E2D3D", tickfont=dict(color="#FFFFFF"),
                          zeroline=False, row=1, col=i)

    st.plotly_chart(fig2, use_container_width=True)
    st.caption("IMHS: non-peak pricing · ZCHS: Mon–Thu or Fri/Sat/Sun tier")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: WSD
# ══════════════════════════════════════════════════════════════════════════════
elif "WSD" in page:
    inject_page_bg("WSD")
    wsd_current, wsd_hist, wsd_spa = load_wsd()
    p = C["WSD"]["p"]; light = C["WSD"]["light"]

    banner("Worldsprings Dallas", "WSD — Dallas, TX  ·  2026 Pricing", "WSD", "💧")

    tab1, tab2, tab3 = st.tabs(["  Current Rates  ", "  Pricing by Day  ", "  Price History  "])

    with tab1:
        col1, col2 = st.columns(2)
        with col1:
            price_table(wsd_current, "Soak Admission 2026", "WSD",
                        "Jan–Feb: same rates · March onward: robe included TBD")
            simple_table([
                ("1-Month Trial (4 day passes)", "$125"),
                ("Weekday Annual (unlimited M-F + 5 guests)", "$948 / yr"),
                ("Anytime Annual (unlimited + 10 guests)", "$1,788 / yr"),
            ], "Memberships", "WSD")
        with col2:
            simple_table([
                ("Robe Rental", "$10"),
                ("Cabana (Mon-Thu)", "$199"),
                ("Cabana (Sat / Holiday)", "$299"),
                ("Day Pass E-Certificate", "$95"),
            ], "Add-ons", "WSD")
            simple_table([
                ("Hot Springs Sound Bath (30 min)", "FREE"),
                ("Guided Contrast Therapy (30 min)", "FREE"),
                ("Stretch, Sauna & Breathwork (30 min)", "FREE"),
                ("Cold Plunge 101 (10 min)", "FREE"),
            ], "Complimentary Classes", "WSD")

    with tab2:
        st.markdown(f"<div style='font-size:16px;font-weight:600;color:{light};margin-bottom:16px'>Price by Day Type — WSD 2026</div>", unsafe_allow_html=True)

        wsd_products = list(wsd_current.keys())  # ["Day Pass", "Evening Pass"]
        wsd_tiers    = ["Mon-Thu", "Fri & Sun", "Saturday"]
        tier_colors  = [p, C["WSD"]["light"], C["WSD"]["dark"]]

        fig = go.Figure()
        for tier, col in zip(wsd_tiers, tier_colors):
            vals = [wsd_current[pr][tier] for pr in wsd_products]
            fig.add_trace(go.Bar(
                name=tier, x=wsd_products, y=vals, marker_color=col,
                text=[f"${v}" for v in vals], textposition="outside",
                textfont=dict(color="#FFFFFF", size=12)
            ))
        layout = base_chart()
        layout.update(barmode="group", height=420, bargap=0.3, bargroupgap=0.08,
            xaxis=dict(tickfont=dict(color="#FFFFFF", size=13)),
            yaxis=dict(tickprefix="$", tickfont=dict(color="#FFFFFF")))
        fig.update_layout(**layout)
        st.plotly_chart(fig, use_container_width=True)

        # Weekend premium KPI cards
        st.markdown(f"<div style='font-size:13px;font-weight:600;color:{p};margin-top:4px;margin-bottom:8px'>Weekend premium by product</div>", unsafe_allow_html=True)
        bg_card = C["WSD"]["dark"] + "33"
        cols_k = st.columns(len(wsd_products) * 2)
        ci = 0
        for pr in wsd_products:
            mth = wsd_current[pr]["Mon-Thu"]
            fri = wsd_current[pr]["Fri & Sun"]
            sat = wsd_current[pr]["Saturday"]
            pct_fri = round((fri - mth) / mth * 100, 1) if mth else 0
            pct_sat = round((sat - mth) / mth * 100, 1) if mth else 0
            for lbl, pct in [(f"{pr} · Fri/Sun", pct_fri), (f"{pr} · Sat", pct_sat)]:
                with cols_k[ci]:
                    st.markdown(f"""<div style='background:{bg_card};border:1px solid {p}30;
                        border-radius:8px;padding:10px;text-align:center'>
                        <div style='font-size:11px;color:#64748B;margin-bottom:4px'>{lbl}</div>
                        <div style='font-size:20px;font-weight:700;color:{p}'>+{pct}%</div>
                    </div>""", unsafe_allow_html=True)
                ci += 1

    with tab3:
        st.markdown(f"<div style='font-size:16px;font-weight:600;color:{light};margin-bottom:4px'>Soak Price History</div>", unsafe_allow_html=True)
        st.caption("WSD launched mid-2024. 2026 rates reflect restructured product names.")

        day_pass_h  = wsd_hist[wsd_hist["Product"].isin(["3 Hour Soak", "Day Pass"])].copy()
        quick_dip_h = wsd_hist[wsd_hist["Product"] == "Quick Dip"].copy()
        evening_h   = wsd_hist[wsd_hist["Product"] == "Evening Pass"].copy()

        fig = go.Figure()
        fig.add_trace(go.Scatter(x=day_pass_h["Period"], y=day_pass_h["Mon-Thu"],
            mode="lines+markers", name="Day Pass (Mon-Thu)",
            line=dict(color=p, width=3), marker=dict(size=9, color=p)))
        fig.add_trace(go.Scatter(x=day_pass_h["Period"], y=day_pass_h["Fri-Sun"],
            mode="lines+markers", name="Day Pass (Fri-Sun)",
            line=dict(color=p, width=3, dash="dot"), marker=dict(size=9, color=p)))
        fig.add_trace(go.Scatter(x=quick_dip_h["Period"], y=quick_dip_h["Mon-Thu"],
            mode="lines+markers", name="Quick Dip (Mon-Thu)",
            line=dict(color=light, width=2), marker=dict(size=8, color=light)))
        fig.add_trace(go.Scatter(x=quick_dip_h["Period"], y=quick_dip_h["Fri-Sun"],
            mode="lines+markers", name="Quick Dip (Fri-Sun)",
            line=dict(color=light, width=2, dash="dot"), marker=dict(size=8, color=light)))
        fig.add_trace(go.Scatter(x=evening_h["Period"], y=evening_h["Mon-Thu"],
            mode="lines+markers", name="Evening Pass (Mon-Thu)",
            line=dict(color=C["WSD"]["dark"], width=2), marker=dict(size=8, color=C["WSD"]["dark"])))
        fig.add_trace(go.Scatter(x=evening_h["Period"], y=evening_h["Fri-Sun"],
            mode="lines+markers", name="Evening Pass (Fri-Sun)",
            line=dict(color=C["WSD"]["dark"], width=2, dash="dot"), marker=dict(size=8, color=C["WSD"]["dark"])))

        chart_layout(fig, "WSD")
        st.plotly_chart(fig, use_container_width=True)

        st.markdown(f"<div style='font-size:14px;font-weight:600;color:{light};margin:8px 0 8px'>Price Table by Period</div>", unsafe_allow_html=True)
        st.dataframe(wsd_hist.set_index("Period"), use_container_width=True)

        # ── Bar chart per period ───────────────────────────────────────────────
        st.markdown(f"<div style='font-size:14px;font-weight:600;color:{light};margin:20px 0 8px'>Price by Period — Key Products</div>", unsafe_allow_html=True)
        day_sel_wsd = st.radio("Price type", ["Mon-Thu", "Fri-Sun"], horizontal=True, key="wsd_hist_tier")

        key_products_wsd = ["Day Pass", "Quick Dip", "Evening Pass"]
        prod_colors_wsd  = [p, C["WSD"]["light"], C["WSD"]["dark"]]
        periods_order_wsd = wsd_hist["Period"].unique().tolist()

        wsd_bar = wsd_hist.copy()
        wsd_bar["Product"] = wsd_bar["Product"].replace("3 Hour Soak", "Day Pass")

        fig_bar = go.Figure()
        for prod, col in zip(key_products_wsd, prod_colors_wsd):
            df_p = wsd_bar[wsd_bar["Product"] == prod].copy()
            df_p["Period"] = pd.Categorical(df_p["Period"], categories=periods_order_wsd, ordered=True)
            df_p = df_p.drop_duplicates("Period").set_index("Period").reindex(periods_order_wsd).reset_index()
            fig_bar.add_trace(go.Bar(
                name=prod, x=df_p["Period"], y=df_p[day_sel_wsd],
                marker_color=col,
                text=df_p[day_sel_wsd].apply(lambda v: f"${v:.0f}" if pd.notna(v) else ""),
                textposition="outside", textfont=dict(color="#FFFFFF", size=11)
            ))
        layout_b = base_chart()
        layout_b.update(barmode="group", height=400, bargap=0.2, bargroupgap=0.08,
            xaxis=dict(tickfont=dict(color="#FFFFFF", size=11)),
            yaxis=dict(tickprefix="$", tickfont=dict(color="#FFFFFF", size=11)),
            legend=dict(font=dict(color="#FFFFFF", size=11)),
        )
        fig_bar.update_layout(**layout_b)
        st.plotly_chart(fig_bar, use_container_width=True)



# ══════════════════════════════════════════════════════════════════════════════
# PAGE: IMHS
# ══════════════════════════════════════════════════════════════════════════════
elif "IMHS" in page:
    inject_page_bg("IMHS")
    imhs_non_peak, imhs_peak, imhs_hist = load_imhs()
    p = C["IMHS"]["p"]; light = C["IMHS"]["light"]

    banner("Iron Mountain Hot Springs", "IMHS — Glenwood Springs, CO  ·  Current Pricing", "IMHS", "🏔️")

    tab1, tab2, tab3 = st.tabs(["  Current Rates  ", "  Pricing by Day  ", "  Price History (2023–Present)  "])

    with tab1:
        col1, col2 = st.columns(2)
        with col1:
            np_data = {d: {"Select 3hr": imhs_non_peak[d]["Select 3hr"],
                           "Select All-Day": imhs_non_peak[d]["Select All-Day"],
                           "Premier 3hr": imhs_non_peak[d]["Premier 3hr"],
                           "Premier All-Day": imhs_non_peak[d]["Premier All-Day"]}
                       for d in DAY_ORDER if d in imhs_non_peak}
            price_table(np_data, "Non-Peak Pricing", "IMHS")
        with col2:
            pk_data = {d: {"Select 3hr": imhs_peak[d]["Select 3hr"],
                           "Select All-Day": imhs_peak[d]["Select All-Day"],
                           "Premier 3hr": imhs_peak[d]["Premier 3hr"],
                           "Premier All-Day": imhs_peak[d]["Premier All-Day"]}
                       for d in DAY_ORDER if d in imhs_peak}
            price_table(pk_data, "Peak / Holiday Pricing", "IMHS")
        simple_table([
            ("Select 3hr", "3-hour timed soak · select pools"),
            ("Select All-Day", "Unlimited duration · select pools"),
            ("Premier 3hr", "3-hour timed soak · all pools"),
            ("Premier All-Day", "Unlimited duration · all pools"),
        ], "Product Tiers", "IMHS")

    with tab2:
        st.markdown(f"<div style='font-size:16px;font-weight:600;color:{light};margin-bottom:16px'>Price by Day of Week — IMHS (current period)</div>", unsafe_allow_html=True)

        days_np   = [d for d in DAY_ORDER if d in imhs_non_peak]
        days_pk   = [d for d in DAY_ORDER if d in imhs_peak]
        metrics_imhs = ["Select 3hr", "Select All-Day", "Premier 3hr", "Premier All-Day"]
        # 4-step green gradient: light → dark
        colors_g4 = ["#A7F3D0", "#4ADE80", "#22C55E", "#15803D"]

        def _imhs_bar(data_src, days, title):
            fig = go.Figure()
            for metric, color in zip(metrics_imhs, colors_g4):
                vals = [data_src.get(d, {}).get(metric, 0) for d in days]
                fig.add_trace(go.Bar(name=metric, x=days, y=vals,
                    marker_color=color,
                    text=[f"${v}" for v in vals], textposition="outside",
                    textfont=dict(color="#FFFFFF", size=10)))
            layout = base_chart()
            layout.update(barmode="group", height=370, bargap=0.2, bargroupgap=0.06,
                title=dict(text=title, font=dict(color="#FFFFFF", size=13), x=0.5),
                xaxis=dict(tickfont=dict(color="#FFFFFF")),
                yaxis=dict(tickprefix="$", tickfont=dict(color="#FFFFFF")),
                legend=dict(font=dict(color="#FFFFFF", size=10)),
                showlegend=True)
            fig.update_layout(**layout)
            return fig

        col_np, col_pk = st.columns(2)
        with col_np:
            st.plotly_chart(_imhs_bar(imhs_non_peak, days_np, "Non-Peak Pricing"), use_container_width=True)
        with col_pk:
            st.plotly_chart(_imhs_bar(imhs_peak, days_pk, "Peak / Holiday Pricing"), use_container_width=True)

        # ── KPIs ──────────────────────────────────────────────────────────────
        def _kpi_card(label, value_str, color, bg):
            return f"""<div style='background:{bg};border:1px solid {color}40;border-radius:8px;
                padding:10px 8px;text-align:center;min-width:0'>
                <div style='font-size:10px;color:#94A3B8;margin-bottom:4px;line-height:1.3'>{label}</div>
                <div style='font-size:18px;font-weight:700;color:{color}'>{value_str}</div>
            </div>"""

        bg_card = C["IMHS"]["dark"] + "33"

        # -- Weekend premium (Non-Peak: Sat vs Mon) --
        np_mon = imhs_non_peak.get("Mon", {})
        np_sat = imhs_non_peak.get("Sat", {})
        st.markdown(f"<div style='font-size:13px;font-weight:600;color:{p};margin:18px 0 8px'>Weekend Premium — Non-Peak (Sat vs Mon)</div>", unsafe_allow_html=True)
        cols_w = st.columns(len(metrics_imhs))
        for i, metric in enumerate(metrics_imhs):
            mon_v = np_mon.get(metric) or 0
            sat_v = np_sat.get(metric) or 0
            pct = round((sat_v - mon_v) / mon_v * 100, 1) if mon_v else 0
            with cols_w[i]:
                st.markdown(_kpi_card(metric, f"+{pct}%", p, bg_card), unsafe_allow_html=True)

        # -- Select vs Premier (Non-Peak Mon) --
        st.markdown(f"<div style='font-size:13px;font-weight:600;color:{p};margin:18px 0 8px'>Premier vs Select Premium — Non-Peak Mon</div>", unsafe_allow_html=True)
        comparisons_sp = [
            ("3hr  ·  Premier vs Select", "Premier 3hr", "Select 3hr"),
            ("All-Day  ·  Premier vs Select", "Premier All-Day", "Select All-Day"),
        ]
        cols_sp = st.columns(2)
        for i, (lbl, higher, lower) in enumerate(comparisons_sp):
            h = np_mon.get(higher) or 0
            l = np_mon.get(lower) or 0
            pct = round((h - l) / l * 100, 1) if l else 0
            with cols_sp[i]:
                st.markdown(_kpi_card(lbl, f"+{pct}%", C["IMHS"]["light"], bg_card), unsafe_allow_html=True)

        # -- Non-Peak vs Peak (Mon and Sat) --
        pk_mon = imhs_peak.get("Mon", {})
        pk_sat = imhs_peak.get("Sat", {})
        st.markdown(f"<div style='font-size:13px;font-weight:600;color:{p};margin:18px 0 8px'>Peak vs Non-Peak Premium</div>", unsafe_allow_html=True)
        ref_days = [("Mon", np_mon, pk_mon), ("Sat", np_sat, pk_sat)]
        for day_lbl, np_d, pk_d in ref_days:
            cols_pk = st.columns(len(metrics_imhs))
            st.markdown(f"<div style='font-size:11px;color:#64748B;margin:6px 0 4px'>{day_lbl}</div>", unsafe_allow_html=True)
            for i, metric in enumerate(metrics_imhs):
                np_v = np_d.get(metric) or 0
                pk_v = pk_d.get(metric) or 0
                pct = round((pk_v - np_v) / np_v * 100, 1) if np_v else 0
                sign = "+" if pct >= 0 else ""
                with cols_pk[i]:
                    st.markdown(_kpi_card(metric, f"{sign}{pct}%", C["IMHS"]["light"], bg_card), unsafe_allow_html=True)

    with tab3:
        st.markdown(f"<div style='font-size:16px;font-weight:600;color:{light};margin-bottom:12px'>Price History — 2022 to Present</div>", unsafe_allow_html=True)

        if not imhs_hist.empty and "Day" in imhs_hist.columns:
            def _parse_dt(s):
                try:
                    p = s.strip().split(".")
                    return pd.Timestamp(f"{p[2]}-{int(p[0]):02d}-{int(p[1]):02d}")
                except Exception:
                    return pd.NaT

            metrics_imhs  = ["Select 3hr", "Select All-Day", "Premier 3hr", "Premier All-Day"]
            colors_np     = ["#A7F3D0", "#4ADE80", "#22C55E", "#15803D"]   # green shades = Non-Peak
            colors_pk     = ["#FCA5A5", "#F87171", "#EF4444", "#B91C1C"]   # red shades  = Peak
            end_ts        = pd.Timestamp("2026-05-31")

            df_h = imhs_hist[imhs_hist["Type"].isin(["Non-Peak", "Peak"])].copy()
            df_h["Date_ts"] = df_h["Date"].apply(_parse_dt)
            df_h = df_h.dropna(subset=["Date_ts"]).sort_values("Date_ts")

            st.markdown("""<style>
            div[data-testid="stRadio"] label {color:#FFFFFF !important; font-weight:600}
            div[data-testid="stRadio"] label p {color:#FFFFFF !important}
            </style>""", unsafe_allow_html=True)

            col_l, col_r = st.columns([1, 2])
            with col_l:
                day_sel = st.radio("Day of week", DAY_ORDER, horizontal=True)
            with col_r:
                prod_sel = st.multiselect("Products", metrics_imhs, default=metrics_imhs)

            df_day = df_h[df_h["Day"] == day_sel].copy()

            # Extend last period to 5.31.26 for both types
            def _extend(df_type):
                df = df_day[df_day["Type"] == df_type].sort_values("Date_ts")
                if df.empty:
                    return df
                last = df.iloc[-1].copy()
                last["Period"] = "5.31.26"; last["Date_ts"] = end_ts
                return pd.concat([df, pd.DataFrame([last])], ignore_index=True)

            np_df = _extend("Non-Peak")
            pk_df = _extend("Peak")

            # ── Trendline: Non-Peak (green) vs Peak (red) ─────────────────────
            fig_trend = go.Figure()
            for metric, cnp, cpk in zip(metrics_imhs, colors_np, colors_pk):
                if metric not in prod_sel:
                    continue
                for df_t, col, label_sfx in [(np_df, cnp, "Non-Peak"), (pk_df, cpk, "Peak")]:
                    df_m = df_t.dropna(subset=[metric])
                    if df_m.empty:
                        continue
                    fig_trend.add_trace(go.Scatter(
                        x=df_m["Date_ts"], y=df_m[metric],
                        name=f"{metric} ({label_sfx})",
                        mode="lines+markers",
                        line=dict(color=col, width=2.5,
                                  dash="dot" if label_sfx == "Peak" else "solid"),
                        marker=dict(size=8, color=col),
                        hovertemplate=f"<b>{metric} – {label_sfx}</b><br>%{{x|%b %Y}}: $%{{y}}<extra></extra>"
                    ))
            layout_t = base_chart()
            layout_t.update(height=440,
                xaxis=dict(tickformat="%b %Y", tickfont=dict(color="#FFFFFF", size=11)),
                yaxis=dict(tickprefix="$", tickfont=dict(color="#FFFFFF", size=11)),
                legend=dict(font=dict(color="#FFFFFF", size=10)),
                title=dict(text=f"Price Trend — {day_sel}  (solid = Non-Peak · dotted = Peak)",
                           font=dict(color="#FFFFFF", size=13), x=0.5),
            )
            fig_trend.update_layout(**layout_t)
            st.plotly_chart(fig_trend, use_container_width=True)

            # ── Scrollable table ──────────────────────────────────────────────
            st.markdown(f"<div style='font-size:14px;font-weight:600;color:{light};margin:20px 0 8px'>Price Table — {day_sel}</div>", unsafe_allow_html=True)
            tbl_rows = []
            for df_t, type_lbl in [(np_df, "Non-Peak"), (pk_df, "Peak")]:
                df_t2 = df_t[df_t["Period"] != "5.31.26"][["Period","Date"] + metrics_imhs].copy()
                df_t2.insert(0, "Type", type_lbl)
                tbl_rows.append(df_t2)
            if tbl_rows:
                tbl = pd.concat(tbl_rows).sort_values(["Date", "Type"])
                for m in metrics_imhs:
                    tbl[m] = tbl[m].apply(lambda v: f"${v:.0f}" if pd.notna(v) else "—")
                st.dataframe(tbl.set_index("Period"), use_container_width=True, height=300)

            # ── Bar chart per period ───────────────────────────────────────────
            st.markdown(f"<div style='font-size:14px;font-weight:600;color:{light};margin:20px 0 8px'>Price by Period — {day_sel}</div>", unsafe_allow_html=True)
            from plotly.subplots import make_subplots
            fig_bar = make_subplots(rows=1, cols=2,
                subplot_titles=["Non-Peak", "Peak"],
                shared_yaxes=True)
            for col_idx, (df_t, cols_t) in enumerate([(np_df, colors_np), (pk_df, colors_pk)], 1):
                df_b = df_t[df_t["Period"] != "5.31.26"].sort_values("Date_ts")
                for metric, col in zip(metrics_imhs, cols_t):
                    fig_bar.add_trace(go.Bar(
                        name=metric, x=df_b["Period"], y=df_b[metric],
                        marker_color=col, showlegend=(col_idx == 1),
                        text=df_b[metric].apply(lambda v: f"${v:.0f}" if pd.notna(v) else ""),
                        textposition="outside", textfont=dict(color="#FFFFFF", size=9)
                    ), row=1, col=col_idx)
            layout_b = base_chart()
            layout_b.update(barmode="group", height=450, bargap=0.15, bargroupgap=0.05,
                xaxis=dict(tickfont=dict(color="#FFFFFF", size=9), tickangle=-30),
                xaxis2=dict(tickfont=dict(color="#FFFFFF", size=9), tickangle=-30),
                yaxis=dict(tickprefix="$", tickfont=dict(color="#FFFFFF", size=11)),
                legend=dict(font=dict(color="#FFFFFF", size=10)),
            )
            fig_bar.update_layout(**layout_b)
            st.plotly_chart(fig_bar, use_container_width=True)
        else:
            st.info("No historical data parsed yet.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: ZCHS
# ══════════════════════════════════════════════════════════════════════════════
elif "ZCHS" in page:
    inject_page_bg("ZCHS")
    zchs_current, zchs_hist = load_zchs()
    p = C["ZCHS"]["p"]; light = C["ZCHS"]["light"]

    banner("Zion Canyon Hot Springs", "ZCHS — Zion Canyon, UT  ·  May 2026 Pricing", "ZCHS", "🏜️")

    tab1, tab2, tab3 = st.tabs(["  Current Rates  ", "  Pricing by Day  ", "  Price History  "])

    with tab1:
        col1, col2 = st.columns(2)
        with col1:
            if zchs_current:
                disp = {pr: {k: v if isinstance(v,(int,float)) else "—" for k,v in prc.items()} for pr,prc in zchs_current.items()}
                price_table(disp, "Soak Pricing (May 2026)", "ZCHS")
        with col2:
            simple_table([("Anytime Annual","$1,499"),("Weekday Annual (M-F)","$920"),("Snowbird (3 months)","$499")], "Memberships", "ZCHS")
            simple_table([("Robe","$10"),("Single Cabana Mon-Thu","$149"),("Single Cabana Fri-Sun","$249"),
                          ("Double Cabana Mon-Thu","$199"),("Double Cabana Fri-Sun","$299")], "Add-ons", "ZCHS")
            simple_table([("Washington County Residents","20% off"),("Hotel Partner Rate","10% off")], "Special Rates", "ZCHS")
        simple_table([("Select (13+)","Standard pools · teens & adults"),
                      ("Select Youth (3–12)","Standard pools · children"),
                      ("Premier (21+)","All pools including 21+ only areas")], "Access Tiers", "ZCHS")

    with tab2:
        st.markdown(f"<div style='font-size:16px;font-weight:600;color:{light};margin-bottom:16px'>Price by Day Type — ZCHS (May 2026)</div>", unsafe_allow_html=True)

        if zchs_current:
            products  = list(zchs_current.keys())
            wkday_vals = [zchs_current[pr].get("Mon-Thu",0) for pr in products]
            wkend_vals = [zchs_current[pr].get("Fri/Sat/Sun",0) for pr in products]
            pct_diff   = [round((w-d)/d*100,1) if d else 0 for d,w in zip(wkday_vals, wkend_vals)]

            fig = go.Figure()
            fig.add_trace(go.Bar(name="Mon–Thu", x=products, y=wkday_vals, marker_color=p,
                text=[f"${v}" for v in wkday_vals], textposition="outside", textfont=dict(color=p)))
            fig.add_trace(go.Bar(name="Fri / Sat / Sun", x=products, y=wkend_vals,
                marker_color=C["ZCHS"]["dark"],
                text=[f"${v}" for v in wkend_vals], textposition="outside", textfont=dict(color=light)))

            layout = base_chart()
            layout.update(barmode="group", height=430, bargap=0.25, bargroupgap=0.08,
                xaxis=dict(tickfont=dict(color="#FFFFFF", size=11), tickangle=-20))
            fig.update_layout(**layout)
            st.plotly_chart(fig, use_container_width=True)

            # weekend premium callouts
            st.markdown(f"<div style='font-size:13px;font-weight:600;color:{p};margin-top:4px;margin-bottom:8px'>Weekend premium by product</div>", unsafe_allow_html=True)
            cols = st.columns(len(products))
            for i, (pr, pct) in enumerate(zip(products, pct_diff)):
                with cols[i]:
                    st.markdown(f"""<div style='background:{C["ZCHS"]["dark"]}33;border:1px solid {p}30;border-radius:8px;
                        padding:10px;text-align:center'>
                        <div style='font-size:11px;color:#64748B;margin-bottom:4px'>{pr}</div>
                        <div style='font-size:20px;font-weight:700;color:{p}'>+{pct}%</div>
                    </div>""", unsafe_allow_html=True)

    with tab3:
        st.markdown(f"<div style='font-size:16px;font-weight:600;color:{light};margin-bottom:12px'>Price History — Since Opening</div>", unsafe_allow_html=True)

        if not zchs_hist.empty:
            date_map = {
                "Feb 2025": pd.Timestamp("2025-02-03"),
                "Jul 2025": pd.Timestamp("2025-07-01"),
                "Nov 2025": pd.Timestamp("2025-11-01"),
            }
            hist_dated = zchs_hist.copy()
            hist_dated["Date"] = hist_dated["Period"].map(date_map)

            # Extend Nov 2025 prices to today (no newer sheet)
            today_ts = pd.Timestamp("2026-05-13")
            nov_rows = hist_dated[hist_dated["Period"] == "Nov 2025"].copy()
            nov_rows["Period"] = "May 2026"
            nov_rows["Date"] = today_ts
            extended = pd.concat([hist_dated, nov_rows], ignore_index=True)

            key_products = ["Quick Dip Select (13+)", "3hr Select (13+)", "Full Day Select (13+)"]
            prod_colors  = [C["ZCHS"]["light"], p, C["ZCHS"]["dark"]]
            periods_order = ["Feb 2025", "Jul 2025", "Nov 2025", "May 2026"]

            day_type = st.radio("Price type", ["Mon-Thu", "Fri/Sat/Sun"], horizontal=True)

            # ── Trendline ─────────────────────────────────────────────────────
            fig_trend = go.Figure()
            for prod, col in zip(key_products, prod_colors):
                df_p = extended[extended["Product"] == prod].sort_values("Date")
                if df_p.empty: continue
                fig_trend.add_trace(go.Scatter(
                    x=df_p["Date"], y=df_p[day_type],
                    name=prod, mode="lines+markers",
                    line=dict(color=col, width=2.5),
                    marker=dict(size=9, color=col),
                    hovertemplate=f"<b>{prod}</b><br>%{{x|%b %Y}}: $%{{y}}<extra></extra>"
                ))
            layout_t = base_chart()
            layout_t.update(height=420,
                xaxis=dict(tickformat="%b %Y", tickfont=dict(color="#FFFFFF", size=11)),
                yaxis=dict(tickprefix="$", tickfont=dict(color="#FFFFFF", size=11)),
                legend=dict(font=dict(color="#FFFFFF", size=11)),
                title=dict(text=f"Price Trend ({day_type})", font=dict(color="#FFFFFF", size=14), x=0.5),
            )
            fig_trend.update_layout(**layout_t)
            st.plotly_chart(fig_trend, use_container_width=True)

            # ── Table ─────────────────────────────────────────────────────────
            st.markdown(f"<div style='font-size:14px;font-weight:600;color:{light};margin:20px 0 8px'>Price Table by Period</div>", unsafe_allow_html=True)
            tbl = extended[extended["Product"].isin(key_products)][["Period","Product","Mon-Thu","Fri/Sat/Sun"]].copy()
            tbl["Period"] = pd.Categorical(tbl["Period"], categories=periods_order, ordered=True)
            tbl = tbl.sort_values(["Period","Product"]).rename(columns={"Period": "Month"})
            tbl["Mon-Thu"]     = tbl["Mon-Thu"].apply(lambda v: f"${v:.0f}" if pd.notna(v) else "—")
            tbl["Fri/Sat/Sun"] = tbl["Fri/Sat/Sun"].apply(lambda v: f"${v:.0f}" if pd.notna(v) else "—")
            st.dataframe(tbl.set_index("Month"), use_container_width=True)

            # ── Bar chart per period ───────────────────────────────────────────
            st.markdown(f"<div style='font-size:14px;font-weight:600;color:{light};margin:20px 0 8px'>Price by Period — Key Products</div>", unsafe_allow_html=True)
            fig_bar = go.Figure()
            for prod, col in zip(key_products, prod_colors):
                df_p = extended[extended["Product"] == prod].copy()
                df_p["Period"] = pd.Categorical(df_p["Period"], categories=periods_order, ordered=True)
                df_p = df_p.drop_duplicates("Period").set_index("Period").reindex(periods_order).reset_index()
                fig_bar.add_trace(go.Bar(
                    name=prod, x=df_p["Period"], y=df_p[day_type],
                    marker_color=col,
                    text=df_p[day_type].apply(lambda v: f"${v:.0f}" if pd.notna(v) else ""),
                    textposition="outside", textfont=dict(color="#FFFFFF", size=11)
                ))
            layout_b = base_chart()
            layout_b.update(barmode="group", height=420, bargap=0.2, bargroupgap=0.08,
                xaxis=dict(tickfont=dict(color="#FFFFFF", size=11)),
                yaxis=dict(tickprefix="$", tickfont=dict(color="#FFFFFF", size=11)),
                legend=dict(font=dict(color="#FFFFFF", size=11)),
            )
            fig_bar.update_layout(**layout_b)
            st.plotly_chart(fig_bar, use_container_width=True)
        else:
            st.info("No historical data parsed yet.")
